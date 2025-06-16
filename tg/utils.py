import re
import random
from datetime import datetime, timedelta
import os
import shutil
from loguru import logger
from functools import wraps

import openpyxl
from openpyxl.drawing.image import Image
import aiofiles
import asyncio

from telethon import TelegramClient
from telethon.errors import AuthKeyDuplicatedError, PhoneNumberInvalidError, FloodWaitError
from telethon.tl.types import MessageEntityUrl, MessageEntityTextUrl, MessageEntityBold, MessageEntityCustomEmoji
from opentele.api import UseCurrentSession
from config import (TELEGRAM_CONFIG, SESSIONS_PATH)

from itertools import cycle
import sqlite3
import csv

# Глобальный словарь для блокировок по файлам
file_locks = {}

def get_file_lock(file_path: str) -> asyncio.Lock:
    """ Возвращает (или создаёт) asyncio.Lock для конкретного файла """
    if file_path not in file_locks:
        file_locks[file_path] = asyncio.Lock()
    return file_locks[file_path]

def is_target_date(post_date_str, target_date_str):
    post_date = datetime.strptime(post_date_str, "%Y-%m-%dT%H:%M:%S+00:00").date()
    target_date = datetime.strptime(target_date_str, "%d.%m.%Y").date()
    return post_date == target_date    

def readFileToStrips(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        return [line.strip() for line in file]

async def async_append_data_to_file(data, path):
    # Убедитесь, что директория существует
    directory = os.path.dirname(path)
    # print(f"Directory: {directory}")
    if not os.path.exists(directory):
        os.makedirs(directory)

    # Теперь безопасно открывайте файл и добавляйте данные
    async with aiofiles.open(path, "a", encoding='utf-8') as file:
        await file.write(data + '\n')
        print('Data was appended to file!')

def append_data_to_file(data, path):
    # Убедитесь, что директория существует
    directory = os.path.dirname(path)
    if not os.path.exists(directory):
        os.makedirs(directory)

    # Теперь безопасно открывайте файл и добавляйте данные
    with open(path, "a", encoding='utf-8') as file:
        file.write(data + '\n')
        # print('Data was appended to file!')

        # ХЗ НАДО НЕТ
        
def convert_to_markdown(message, entities):
    markdown_text = message
    offset_correction = 0
    
    for entity in sorted(entities, key=lambda x: x.offset):  # Отсортируем entities по offset, чтобы корректно обрабатывать смещения
        start = entity.offset + offset_correction
        end = start + entity.length
        
        if isinstance(entity, MessageEntityBold):
            markdown_text = markdown_text[:start] + '**' + markdown_text[start:end] + '**' + markdown_text[end:]
            offset_correction += 4  # Добавлены 4 символа: две звездочки в начале и две в конце
            
        elif isinstance(entity, MessageEntityTextUrl):
            link_text = markdown_text[start:end]
            markdown_link = f"[{link_text}]({entity.url})"
            markdown_text = markdown_text[:start] + markdown_link + markdown_text[end:]
            offset_correction += len(markdown_link) - len(link_text)  # Учитываем изменение длины строки из-за добавления Markdown
            
        elif isinstance(entity, MessageEntityCustomEmoji):
            # Если у вас есть способ преобразовать custom emoji в markdown, добавьте его здесь
            pass
        
    return markdown_text
        # ХЗ НАДО НЕТ

async def set_profile(client, name, avatar_file):
    await client.update_profile(first_name=name)
    await client.upload_profile_photo(file=avatar_file)

def calculate_delay(account_index, total_accounts):
    # 60% аккаунтов вступают в первые 10 минут
    if account_index < total_accounts * 0.6:
        return random.uniform(0, 10 * 60)
    # 30% аккаунтов в следующие 20 минут
    elif account_index < total_accounts * 0.9:
        return random.uniform(10 * 60, 30 * 60)
    # 10% аккаунтов в течении следующего часа
    else:
        return random.uniform(30 * 60, 90 * 60)


def delete_data(account):
    # Удаление данных об аккаунте из таблицы accounts
    with sqlite3.connect('telegram_bot.db') as conn:
        c = conn.cursor()
        c.execute("DELETE FROM accounts WHERE account = ?", (account,))
        conn.commit()
    # Удаление данных о сессии из таблицы sessions
    with sqlite3.connect('sessions.db') as conn:
        c = conn.cursor()
        c.execute("DELETE FROM sessions WHERE session_name = ?", (account,))
        conn.commit()
    
    # Удаление файла сессии
    # Копирование файла
    shutil.copy("sessions/" + session, "cash/trashAccs/" + session)

    # Удаление исходного файла
    os.remove("sessions/" + session)

def initialize_database(sessions, channels):
    conn = sqlite3.connect('telegram_bot.db')
    c = conn.cursor()
    
    # Убедитесь, что таблица accounts создана
    c.execute('''CREATE TABLE IF NOT EXISTS accounts (account TEXT, channel TEXT, UNIQUE(channel))''')
    conn.commit()

    # Распределение каналов по сессиям
    for i, channel in enumerate(channels):
        session = sessions[i % len(sessions)]
        
        # Проверяем, существует ли уже канал в любой записи
        c.execute("SELECT * FROM accounts WHERE channel = ?", (channel,))
        exists = c.fetchone()
        
        # Если канал не существует, добавляем его
        if not exists:
            try:
                c.execute("INSERT INTO accounts (account, channel) VALUES (?, ?)", (session, channel))
                print(f"Добавлен канал {channel} для сессии {session}.")
            except sqlite3.IntegrityError:
                print(f"Не удалось добавить канал {channel} из-за конфликта уникальности.")
        else:
            print(f"Канал {channel} уже существует в базе данных. Пропускаем.")

    conn.commit()
    conn.close()

def create_tables():
    with get_db_connection('telegram_bot.db') as conn:
        conn.execute('''CREATE TABLE IF NOT EXISTS accounts (
                        account TEXT, 
                        channel TEXT, 
                        invite_requests INTEGER DEFAULT 0
                    )''')
        # Создаём уникальный индекс для столбца channel
        conn.execute('''CREATE UNIQUE INDEX IF NOT EXISTS idx_channel_unique ON accounts(channel)''')
        conn.commit()
    
    with get_db_connection('sessions.db') as conn:
        conn.execute('''CREATE TABLE IF NOT EXISTS sessions (
                        session_name TEXT, 
                        start_time TEXT, 
                        request_count INTEGER
                    )''')
        conn.commit()


def move_dead_session(session: str, sessions_path: str, dead_sessions_folder: str = "sessionsDied"):
    """
    Перемещает файл сессии в папку 'sessionsDied'.

    :param session: Имя файла сессии (без пути)
    :param sessions_path: Директория, где хранятся активные сессии
    :param dead_sessions_folder: Директория, куда перемещаются мёртвые сессии
    """
    os.makedirs(dead_sessions_folder, exist_ok=True)

    session_file = os.path.join(sessions_path, session)
    destination_file = os.path.join(dead_sessions_folder, session)

    logger.info(f"Попытка переместить файл сессии: {session_file} -> {destination_file}")

    if os.path.exists(session_file):
        try:
            shutil.move(session_file, destination_file)
            logger.info(f"Файл сессии {session} перемещён в {dead_sessions_folder}")
        except Exception as e:
            logger.error(f"Ошибка перемещения файла сессии {session}: {e}")
    else:
        logger.warning(f"Файл {session_file} не найден. Возможно, он уже удалён.")

def read_excel_data(excel_file):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]  # Заголовки
    excel_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):  # Проверка на пустую строку
            excel_data.append(dict(zip(headers, row)))
    return excel_data, workbook, sheet  # Возвращаем workbook и sheet для дальнейших изменений


def create_excel_file_with_avatars(avatars_folder):
    excel_file = 'futureAccs.xlsx'
    if not os.path.exists(excel_file):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Добавляем заголовки столбцов
        sheet.append(["Сессия", "Путь к изображению", "Изображение", "Имя", "Фамилия", "Имя пользователя", "Описание профиля", "Пол", "Прокси"])
        workbook.save(excel_file)

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    files = [f for f in os.listdir(avatars_folder) if os.path.isfile(os.path.join(avatars_folder, f)) and f.lower().endswith(('.png', '.jpg', '.jpeg'))]
    file_count = 0

    for file in files:
        original_img_path = os.path.join(avatars_folder, file)
        resized_img_path = os.path.join(avatars_folder, "resized_" + file)

        # Изменение размера изображения с помощью Pillow
        with PilImage.open(original_img_path) as img:
            img.thumbnail((250, 250), PilImage.Resampling.LANCZOS)  # Изменяем размер до 250x250
            img.save(resized_img_path)  # Сохраняем измененное изображение

        sheet.append([None, original_img_path, None, None, None, None, None, None])
        row = sheet.max_row; col = 3  # Указываем столбец для изображения

        # Устанавливаем высоту строки
        sheet.row_dimensions[row].height = 333

        # Добавляем измененное изображение в Excel
        img = Image(resized_img_path)
        cell_address = openpyxl.utils.get_column_letter(col) + str(row)
        img.anchor = cell_address
        sheet.add_image(img)

        file_count += 1
        if file_count % 4 == 0:
            sheet.append([])

    workbook.save(excel_file)


def handle_flood_wait():
    def decorator(func):
        async def wrapper(*args, **kwargs):
            while True:  # бесконечный цикл для повторных попыток
                try:
                    return await func(*args, **kwargs)
                except FloodWaitError as e:
                    logger.error(f"FloodWaitError: ожидание {e.seconds} секунд")
                    await asyncio.sleep(e.seconds + 85)
        return wrapper
    return decorator

# Функция для обработки неавторизованных сессий (работает с таблицей sessions в telegram_bot.db)
async def handle_un_auth_error(session: str, sessions_list: list):
    logger.error(f"Ошибка авторизации сессии {session}")
    with sqlite3.connect("sessions.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM sessions WHERE session_name = ?", (session,))
        session_info = cursor.fetchone()

    if session_info:
        session_name, start_time, request_count = session_info
        duration = datetime.now() - datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S")

        parts = session.split("_")
        seller = parts[0]
        country = parts[1] if len(parts) > 1 else ""
        index = parts[2] if len(parts) > 2 else ""
        price = parts[-1]

        with open("session_info.csv", "a", newline="", encoding="utf-8") as csvfile:
            fieldnames = ["seller", "country", "index", "price", "duration", "request_count"]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writerow({
                "seller": seller,
                "country": country,
                "index": index,
                "price": price,
                "duration": duration.total_seconds() / 3600,
                "request_count": request_count,
            })

    with sqlite3.connect("telegram_bot.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT channel FROM accounts WHERE account = ?", (session,))
        channels_for_session = [row[0] for row in cursor.fetchall()]
        cursor.execute("DELETE FROM accounts WHERE account = ?", (session,))
        conn.commit()

    if not channels_for_session:
        logger.info(f"Для сессии {session} не найдено каналов для перераспределения.")
        return

    logger.info(f"Сессия {session} заблокирована. Перераспределяем {len(channels_for_session)} каналов.")
    active_sessions = [s for s in sessions_list if s != session]
    if not active_sessions:
        logger.error("Нет доступных сессий для перераспределения каналов!")
        return

    with sqlite3.connect("telegram_bot.db") as conn:
        cursor = conn.cursor()
        conn.execute("BEGIN")
        for i, channel in enumerate(channels_for_session):
            target_session = active_sessions[i % len(active_sessions)]
            try:
                cursor.execute("INSERT INTO accounts (account, channel) VALUES (?, ?)", (target_session, channel))
            except sqlite3.IntegrityError:
                logger.warning(f"Канал {channel} уже распределён, пропускаем")
        conn.commit()
    logger.info("Каналы успешно перераспределены между оставшимися сессиями.")

def check_authorization():
    def decorator(func):
        @wraps(func)
        async def wrapper(self, session: str, *args, **kwargs):
            client = TelegramClient(os.path.join("sessions", session), **TELEGRAM_CONFIG)
            try:
                await client.connect()
                logger.info(f"{client.session.filename} - Запрос к API: connect")
                
                if not await client.is_user_authorized():
                    logger.error(f"{client.session.filename} - не авторизован")
                    await handle_un_auth_error(session, self.sessions)
                    move_dead_session(session, SESSIONS_PATH)
                    return None
                
                logger.info(f"{client.session.filename} - Авторизован")
                await asyncio.sleep(25)  # Сохраняем задержку
                
                # Передаем клиент в оригинальную функцию вместо session
                return await func(self, client, *args, **kwargs)
                
            except (AuthKeyDuplicatedError, PhoneNumberInvalidError) as e:
                error_type = "дублирование ключа" if isinstance(e, AuthKeyDuplicatedError) else "неверный номер"
                logger.error(f"{client.session.filename} - Ошибка авторизации: {error_type}")
                await handle_un_auth_error(session, self.sessions)
                move_dead_session(session, SESSIONS_PATH)
                return None
        return wrapper
    return decorator


async def convertToTDATA(session: str):
    """
    Загружает клиента Telethon из файла сессии и конвертирует его в формат TDesktop.
    
    Args:
        session: Имя файла сессии (без пути)
    """
    # Загружаем клиента Telethon из файла сессии.
    # Если не указывать api_id и api_hash, будет использован API Telegram Desktop по умолчанию.
    client = TelegramClient(os.path.join("sessions", session), **TELEGRAM_CONFIG)
    await client.connect()

    # Режим UseCurrentSession – используем существующую сессию Telethon.
    tdesk = await client.ToTDesktop(flag=UseCurrentSession)

    # Сохраняем TDesktop-сессию в папку "tdata"
    tdesk.SaveTData("tdata")
    logger.info("Сессия успешно конвертирована в tdata!")

    await client.disconnect()

class TelegramChannel:
    """
    Класс для работы с каналом Telegram.
    
    Attributes:
        channel_name (str): Имя канала Telegram
    """
    def __init__(self, channel_name: str):
        self.channel_name = channel_name

    def is_private(self) -> bool:
        """
        Проверяет, является ли канал приватным.
        
        Returns:
            bool: True, если канал приватный, иначе False.
        """
        return "t.me/joinchat/" in self.channel_name

    def formatted(self) -> str:
        """
        Возвращает форматированное имя канала.
        
        Returns:
            str: Форматированное имя канала.
        """
        return (self.channel_name.split("t.me/joinchat/")[-1] if self.is_private() else 
                self.channel_name.split("@")[-1] if "@" in self.channel_name else
                self.channel_name.split("+")[-1] if "+" in self.channel_name else
                self.channel_name.split("t.me/")[-1])
